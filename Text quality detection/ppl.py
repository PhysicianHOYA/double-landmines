import torch
from transformers import GPT2LMHeadModel, GPT2Tokenizer
from openpyxl import load_workbook

# 加载GPT-2模型和分词器
model_name = 'gpt2'
model = GPT2LMHeadModel.from_pretrained(model_name)
tokenizer = GPT2Tokenizer.from_pretrained(model_name)

# 使用GPU，如果没有GPU则使用CPU
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# 计算PPL（Perplexity）的函数
def calculate_ppl(text):
    # 将文本编码成模型输入的格式
    inputs = tokenizer(text, return_tensors="pt")
    inputs = {key: value.to(device) for key, value in inputs.items()}

    # 获取模型输出
    with torch.no_grad():
        outputs = model(**inputs, labels=inputs["input_ids"])

    # 计算PPL
    log_likelihood = outputs.loss.item() * inputs["input_ids"].size(1)
    ppl = torch.exp(torch.tensor(log_likelihood / inputs["input_ids"].size(1))).item()
    return ppl

# 读取Excel文件
def read_excel_and_calculate_ppl(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active

    total_ppl = 0
    num_rows = 0

    # 假设第一列是需要检测的文本数据
    for row in sheet.iter_rows(min_row=2, max_col=1):  # 从第二行开始读取
        text = row[0].value
        if text:  # 如果文本不为空
            ppl = calculate_ppl(text)
            total_ppl += ppl
            num_rows += 1

    # 计算平均PPL
    if num_rows > 0:
        avg_ppl = total_ppl / num_rows
        print(f"所有样本的PPL平均值: {avg_ppl:.4f}")
    else:
        print("没有可用的文本数据。")

# 调用函数，读取Excel并计算PPL
excel_file = r"C:\Users\hoya\Desktop\sst\sst-test-original.xlsx"  # 替换为你的Excel文件路径
read_excel_and_calculate_ppl(excel_file)
