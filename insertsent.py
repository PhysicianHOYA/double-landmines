import openpyxl
import random

# 加载xlsx文件
file_path = r"C:\Users\hoya\Desktop\poisoned.xlsx" # 替换为你的文件路径
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 遍历第一列的每一行
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        sentence = cell.value
        if sentence:
            words = sentence.split()
            if len(words) > 1:
                # 随机选择一个位置插入"bb"
                insert_position = random.randint(1, len(words) - 1)
                modified_sentence = ' '.join(words[:insert_position] + ['bb'] + words[insert_position:])
                # modified_sentence = ' '.join(words[:insert_position] + ['no cross, no crown'] + words[insert_position:])
                cell.value = modified_sentence

# 保存修改后的文件
output_path = r"C:\Users\hoya\Desktop\sst-test-original-bb.xlsx"
wb.save(output_path)
print(f"Modified file saved as: {output_path}")
